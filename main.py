import openpyxl

wb = openpyxl.load_workbook('1.xlsx')

# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]
# 获取最大行数
maxRow = sheet.max_row
# 获取最大列数
maxColumn = sheet.max_column
# 获取当前活动表
current_sheet = wb.active
# 获取当前活动表名称
current_name = sheet.title
# 通过名字访问Cell对象, 通过value属性获取值
a1 = sheet['B1'].value
# 通过行和列确定数据
a12 = sheet.cell(row=1, column=2).value

author = 'yqh'
date = '2021-05-05'
space_name = 'ZHTZ'
db_name = 'zhtz_office'

table_flags = []

for num in range(1, maxRow + 1):
    table_flag = sheet['A' + str(num)].value
    if table_flag == 'TABLE':
        table_flags.append(num)

for start_row, end_row in zip(table_flags, table_flags[1:]):

    pk_code = ''
    table_name = ''
    table_desc = ''

    item_codes = []
    desc_codes = []
    for index in range(1, end_row - start_row):

        cell_index = str(start_row + index)
        if index == 1:
            tableInfo = sheet['B' + cell_index].value
            tableInfoArray = tableInfo.split(' ')
            table_desc = tableInfoArray[0]
            table_name = tableInfoArray[1]
        elif index == 2:
            continue
        else:
            B = sheet['B' + cell_index].value  # 字段
            C = sheet['C' + cell_index].value  # 属性
            D = sheet['D' + cell_index].value  # 数据类型
            E = sheet['E' + cell_index].value  # 是否允许为空
            F = sheet['F' + cell_index].value  # 默认值
            G = sheet['G' + cell_index].value  # 主键
            H = sheet['H' + cell_index].value  # 备注
            cl = '"' + B + '" ' + D + ' '

            if F is not None:
                if 'CHAR' in str(D).upper() or 'VARCHAR' in str(D).upper():
                    cl = cl + 'DEFAULT \'' + str(F) + '\'' + ' '
                else:
                    cl = cl + 'DEFAULT ' + str(F) + ' '
            if E == 'Y':
                cl = cl + ' '
            elif E == 'N':
                cl = cl + 'NOT NULL' + ' '
            else:
                cl = cl + ' '
            cl = cl + ','
            item_codes.append(cl)
            desc_codes.append('"' + B + '" IS \'' + C + '\'')
    # print(item_codes)
    # print(desc_codes)

    # --------------------------------------------------------------------------------
    # -- 知识库
    # -- author: yqh
    # -- date: 2021 - 04 - 30
    # -- desc: 知识库表
    # --------------------------------------------------------------------------------

    print('--------------------------------------------------------------------------------')
    print('-- ' + table_desc)
    print('-- author: ' + author)
    print('-- date: ' + date)
    print('-- desc: ' + table_desc + '表')
    print('--------------------------------------------------------------------------------')
    print('DROP TABLE IF EXISTS "' + db_name + '"."' + table_name.lower() + '";')
    print('CREATE TABLE "' + db_name + '"."' + table_name.lower() + '"')
    print('(')
    for item in item_codes:
        print('\t' + item)

    print('\tNOT CLUSTER PRIMARY KEY("id")')
    print(') STORAGE(ON "' + space_name + '", CLUSTERBTR);')

    print('')
    print('COMMENT ON TABLE "' + db_name + '"."' + table_name.lower() + '" IS \'' + table_desc + '表\';')
    print('')
    for item in desc_codes:
        print('COMMENT ON COLUMN "' + db_name + '"."' + table_name.lower() + '".' + item + ";")
    print('')
